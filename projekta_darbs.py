import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

import requests
import bs4

import time

import openpyxl
from openpyxl import workbook

import re

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service = service, options = option)

preces_data = []

#nosaka cik preču un kādas preces tiks meklētas
print("Cik preces tiks meklētas: ")
n = input()

print("Kādas preces tiks meklētas: ")
for _ in range(int(n)):
    x = input()
    preces_data.append(x)

#tiek atvērts 1a.lv
url = "https://www.1a.lv"
driver.get(url)
time.sleep(2)

#tiek meklētas attiecīgās preces 1a.lv un dodas uz to rezultātu mājaslapu
for line in preces_data:

    find = driver.find_element(By.ID, "q")
    find.clear()
    find.send_keys(line)

    button = driver.find_element(By.CLASS_NAME, "main-search__submit")
    button.click()
    time.sleep(2)

    #tiek atrasti top 3 populārāko preču parametri, nosaukumi, cenas un tie tiek ierakstīti csv failā, lai nepazaudētu informāciju un lai būtu vieglāk rakstīt kodu pa daļām
    ##parametri
    pop_preces = 0
    f = open("parametri_"+ line + ".csv", "a", encoding = "utf-8")

    lapas_saturs = bs4.BeautifulSoup(driver.page_source, "html.parser")
    parametri = lapas_saturs.find_all(class_ = "ks-new-product-attributes")

    for i in parametri:
        if pop_preces < 5:
            pop_preces += 1
            f.write(str(i.text.strip()) + " " + "\n")
        else:
            break
    f.close()

    #--------------------------------------------------
    ##nosaukumi
    pop_preces = 0
    f = open("nosaukumi_" + line + ".csv", "a", encoding = "utf-8")

    lapas_saturs = bs4.BeautifulSoup(driver.page_source, "html.parser")
    nosaukums = lapas_saturs.find_all(class_ = "ks-new-product-name")

    for i in nosaukums:
        if pop_preces< 5:
            pop_preces += 1
            f.write(str(i.text.strip()) + " " + "\n")
        else:
            break
    f.close()

    #--------------------------------------------------
    ##cenas
    pop_preces = 0
    f = open("cenas_" + line + ".csv", "a", encoding = "utf-8")

    lapas_saturs = bs4.BeautifulSoup(driver.page_source, "html.parser")
    cena = lapas_saturs.find_all(class_ = "ks-new-product-price")

    for i in cena:
        if pop_preces < 5:
            pop_preces += 1
            f.write(str(i.text.strip()) + " " + "\n")
        else:
            break
    f.close()

#dati no csv failiem tiek pārvietoti uz excel failu attiecīgi izvelētās excel šūnās
parametri_data = []
cenas_data = []
nosaukumi_data = []

wb = openpyxl.Workbook()
jauns_nosaukums = "rezultats"
page = wb.active
page.title = jauns_nosaukums

for line in preces_data:
    page = wb.create_sheet(line)
    page['A1'] = "Nosaukums"
    page['E1'] = "Cena (EUR)"

##parametri
###videokarte
page = wb["videokarte"]

with open("parametri_videokarte.csv", "r", encoding = "utf-8") as f:
    for line in f:
            skaitlis = re.findall(r'\b\d+\b', line)
            parametri_data.append(skaitlis)

for i in range(len(parametri_data)):
    frekvence = parametri_data[i][0]
    page["B" + str(i + 2)].value = int(frekvence)

    atmina = parametri_data[i][1]
    page["C" + str(i + 2)].value = int(atmina)

parametri_data.clear()

#-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-
###procesors

page = wb["procesors"]
with open("parametri_procesors.csv", "r", encoding = "utf-8") as f:
    for line in f:
        skaitlis = re.findall(r'(\b\d+.)', line)
        parametri_data.append(skaitlis)

for i in range(len(parametri_data)):
     frekvence = parametri_data[i][0] + parametri_data[i][1]
     page["B" + str(i + 2)].value = float(frekvence)

     kodoli = parametri_data[i][2].replace("K", "")
     page["C" + str(i + 2)].value = int(kodoli)

     kesatmina = parametri_data[i][3]
     page["D" + str(i + 2)].value = int(kesatmina)

parametri_data.clear()

#-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-.-
###cietais disks

page = wb["cietais disks"]
with open("parametri_cietais disks.csv", "r", encoding = "utf=8") as f:
     for line in f:
          skaitlis = re.findall(r'\b\d+\b', line)
          parametri_data.append(skaitlis)

for i in range(len(parametri_data)):
     ietilpiba = parametri_data[i][0]
     if int(ietilpiba) == 1:
         page["B" + str(i + 2)].value = int(ietilpiba) * 1024
     else:
         page["B" + str(i + 2)].value = int(ietilpiba)

     rakst_v = parametri_data[i][1]
     page["C" + str(i + 2)].value = int(rakst_v)

     las_v = parametri_data[i][2]
     page["D" + str(i + 2)].value = int(las_v)

##nosaukumi
for line in preces_data:
    page = wb[line]
    with open("nosaukumi_" + line + ".csv", "r", encoding = "utf-8") as f:
        for line in f:
            nosaukumi = line.split(",")[0]
            nosaukumi_data.append(nosaukumi)
            
            for i in range(len(nosaukumi_data)):
                nosaukums = nosaukumi_data[i]
                page["A" + str(i + 2)].value = nosaukums

    nosaukumi_data.clear()

##cenas
for line in preces_data:
    page = wb[line]
    with open("cenas_" + line + ".csv", "r", encoding = "utf-8") as f:
        for line in f:
            cenas = re.findall(r'\b\d+\b', line)
            cenas_data.append(cenas)

            for i in range(len(cenas_data)):
                cena = cenas_data[i][0] + "." + cenas_data[i][1]
                page["E" + str(i + 2)].value = float(cena)

    cenas_data.clear()

#diemžēl koda izpildes laikā radušies duplikāti, jo 1a.lv visu preču parametri, cenas un nosaukumi tiek duplicēti, tāpēc tiks izdzēsta otrā rinda katrā excel preču lapā
izdzest_1 = 2
izdzest_2 = 3

for line in preces_data:
    page = wb[line]
    page.delete_rows(izdzest_1)
    page.delete_rows(izdzest_2)

#tiek aizpildīts excel šūnas noteiktās lapās, lai ierakstītā dati būtu saprotami un viegli uztverami
videokarte = wb["videokarte"]
videokarte['B1'] = "Frekvence (MHz)"
videokarte['C1'] = "Atmiņa (GB)"
videokarte['D1'] = "---------------"

procesors = wb['procesors']
procesors['B1'] = "Frekvence (GHz)"
procesors['C1'] = "Kodolu skaits"
procesors['D1'] = "Kešatmiņa (MB)"

cietais_disks = wb['cietais disks']
cietais_disks['B1'] = "Atmiņa (GB)"
cietais_disks['C1'] = "Rakstīšanas ātrums (Mb/s)"
cietais_disks['D1'] = "Lasīšanas ātrums (Mb/s)"

page = wb["rezultats"]
page["C1"] = "Atmiņa (GB)"
page["D1"] = "Kešatmiņa (MB)"
page["E1"] = "Kodoli"
page["F1"] = "Rakstīšanas ātrums (Mb/s)"
page["G1"] = "Lasīšanas ātrums (Mb/s)"
page["H1"] = "Cena (EUR)"

#veiktspējīgākās/ietilpīgākās preces tiek ievietotas rezūltātu lapā
##videokarte
page = wb["videokarte"]
max_row = page.max_row

best_atmina = 0
best_cena = 0
best_nosaukums = ""

for i in range(2,max_row + 1):
    page = wb["videokarte"]

    if i < max_row - 1:
        atmina = int(page["C" + str(i)].value)
        nosaukums = page["A" + str(i)].value
        cena = float(page["E" + str(i)].value)

        if atmina > best_atmina:
            best_atmina = atmina
            best_cena = cena
            best_nosaukums = nosaukums
        elif atmina == best_atmina and cena < best_cena:
            best_cena = cena
            best_nosaukums = nosaukums

page = wb["rezultats"]
page["B2"].value = best_nosaukums
page["C2"].value = best_atmina
page["D2"].value = "---------------"
page["E2"].value = "---------------"
page["F2"].value = "---------------"
page["G2"].value = "---------------"
page["H2"].value = best_cena


#------------------------------------------------------
##procesors
page = wb["procesors"]

max_row = page.max_row

for i in range(2, max_row + 1):
    page = wb["procesors"]

    best_kodoli = 0
    best_kesatmina = 0
    best_cena = 0
    best_nosaukums = ""

    if i < max_row - 1:
        kodoli = int(page["C" + str(i)].value)
        kesatmina = int(page["D" + str(i)].value)
        nosaukums = page["A" + str(i)].value
        cena = float(page["E" + str(i)].value)

        if kodoli > best_kodoli:
            best_kodoli = kodoli
            best_kesatmina = kesatmina
            best_cena = cena
            best_nosaukums = nosaukums
        elif kodoli == best_kodoli and kesatmina > best_kesatmina:
            best_kesatmina = kesatmina
            best_cena = cena
            best_nosaukums = nosaukums
        elif kodoli == best_kodoli and kesatmina == best_kesatmina and cena < best_cena:
            best_cena = cena
            best_nosaukums = nosaukums

page = wb["rezultats"]
page["B3"].value = best_nosaukums
page["C3"].value = "---------------"
page["D3"].value = best_kesatmina
page["E3"].value = best_kodoli
page["F3"].value = "---------------"
page["G3"].value = "---------------"
page["H3"].value = best_cena

#------------------------------------------------------
##procesors
page = wb["cietais disks"]
max_row = page.max_row

best_atmina = 0
best_las_v = 0
best_rakst_v = 0
best_cena = 0
best_nosaukums = ""

for i in range(2, max_row + 1):
    page = wb["cietais disks"]

    if i < max_row - 1:
        nosaukums = page["A" + str(i)].value
        atmina = int(page["B" + str(i)].value)
        las_v = int(page["D" + str(i)].value)
        rakst_v = int(page["C" + str(i)].value)
        cena = float(page["E" + str(i)].value)

        if atmina > best_atmina:
            best_atmina = atmina
            best_las_v = las_v
            best_rakst_v = rakst_v
            best_cena = cena
            best_nosaukums = nosaukums
        elif atmina == best_atmina and las_v > best_las_v:
            best_las_v = las_v
            best_rakst_v = rakst_v
            best_cena = cena
            best_nosaukums = nosaukums
        elif atmina == best_atmina and las_v == best_las_v and rakst_v > best_rakst_v:
            best_rakst_v = rakst_v
            best_cena = cena
            best_nosaukums = nosaukums
        elif atmina == best_atmina and las_v == best_las_v and rakst_v == best_rakst_v and cena < best_cena:
            best_cena = cena
            best_nosaukums = nosaukums

page = wb["rezultats"]
page["B4"].value = best_nosaukums
page["C4"].value = best_atmina
page["D4"].value = "---------------"
page["E4"].value = "---------------"
page["F4"].value = best_rakst_v
page["G4"].value = best_las_v
page["H4"].value = best_cena

cenas = []

for i in range(2, max_row + 1):
    cena = page["H" + str(i + 2)].value
    cenas.append(cena)
    summa = sum(cenas)

page["H" + str(max_row + 1)].value = "Kopā: " + summa

wb.save("rezultats_0.xlsx")
wb.close()
#Aldis Gulbis 201RMC132